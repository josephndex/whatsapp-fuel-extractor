#!/usr/bin/env python3
"""
WhatsApp Fuel Extractor - Weekly Summary Generator

Generates weekly fuel consumption statistics and formats them for WhatsApp.
Statistics include:
- Total liters consumed
- Total amount spent
- Number of cars fueled
- Top consumer (most liters)
- Most distance (highest odometer increase)
- Worst efficiency (lowest km/liter ratio)
"""

import os
import json
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

from openpyxl import load_workbook

# Setup logging with rotation
LOG_FILE = Path(__file__).parent.parent / 'weekly_summary.log'
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)

file_handler = RotatingFileHandler(
    LOG_FILE,
    maxBytes=5 * 1024 * 1024,  # 5 MB
    backupCount=3,
    encoding='utf-8'
)
file_handler.setFormatter(log_formatter)

logging.basicConfig(level=logging.INFO, handlers=[console_handler, file_handler])
logger = logging.getLogger(__name__)

# Paths
ROOT_DIR = Path(__file__).parent.parent
CONFIG_PATH = ROOT_DIR / 'config.json'
SUMMARY_OUTPUT_PATH = ROOT_DIR / 'data' / 'weekly_summary.json'


def load_config() -> Dict:
    """Load configuration from config.json."""
    try:
        with open(CONFIG_PATH, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading config: {e}")
        return {}


def get_excel_path() -> Optional[Path]:
    """Get the path to the Excel file from config."""
    config = load_config()
    output_folder = config.get('output', {}).get('excelFolder', './data/output')
    excel_filename = config.get('output', {}).get('excelFileName', 'fuel_records.xlsx')
    
    if output_folder.startswith('./'):
        output_folder = ROOT_DIR / output_folder[2:]
    else:
        output_folder = Path(output_folder)
    
    return output_folder / excel_filename


def load_weekly_data(days: int = 7) -> List[Dict]:
    """Load fuel records from the last N days."""
    excel_path = get_excel_path()
    
    if not excel_path or not excel_path.exists():
        logger.warning(f"Excel file not found: {excel_path}")
        return []
    
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        
        records = []
        cutoff_date = datetime.now() - timedelta(days=days)
        
        # Get column headers from first row
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            record = dict(zip(headers, row))
            
            # Parse datetime to check if within range
            datetime_str = record.get('DATETIME', '')
            if datetime_str:
                try:
                    # Handle format: YYYY-MM-DD-HH-MM
                    record_date = datetime.strptime(datetime_str, '%Y-%m-%d-%H-%M')
                    if record_date >= cutoff_date:
                        records.append(record)
                except ValueError:
                    # Try alternative formats
                    try:
                        record_date = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
                        if record_date.replace(tzinfo=None) >= cutoff_date:
                            records.append(record)
                    except:
                        # Include record if date can't be parsed
                        records.append(record)
        
        wb.close()
        logger.info(f"Loaded {len(records)} records from the last {days} days")
        return records
        
    except Exception as e:
        logger.error(f"Error loading Excel data: {e}")
        return []


def calculate_statistics(records: List[Dict], days: int = 7) -> Dict:
    """Calculate statistics from fuel records for the given period."""
    if not records:
        return None
    
    # Initialize stats
    total_liters = 0.0
    total_amount = 0.0
    cars_fueled = set()
    drivers = set()
    departments = set()
    fuel_types = defaultdict(float)  # fuel_type -> liters
    
    # Track per-car and per-driver consumption
    car_liters = defaultdict(float)
    car_amount = defaultdict(float)
    car_odometer_readings = defaultdict(list)
    car_fuel_count = defaultdict(int)
    driver_liters = defaultdict(float)
    driver_amount = defaultdict(float)
    dept_liters = defaultdict(float)
    dept_amount = defaultdict(float)
    
    # Track daily breakdown
    daily_liters = defaultdict(float)
    daily_amount = defaultdict(float)
    daily_count = defaultdict(int)
    
    for record in records:
        # Get car plate
        car = str(record.get('CAR', '')).strip().upper()
        if not car:
            continue
        
        cars_fueled.add(car)
        car_fuel_count[car] += 1
        
        # Get driver
        driver = str(record.get('DRIVER', '')).strip().title()
        if driver:
            drivers.add(driver)
        
        # Get department
        dept = str(record.get('DEPARTMENT', '')).strip().upper()
        if dept:
            departments.add(dept)
        
        # Get liters
        liters = record.get('LITERS')
        if liters:
            try:
                liters_val = float(str(liters).replace(',', ''))
                total_liters += liters_val
                car_liters[car] += liters_val
                if driver:
                    driver_liters[driver] += liters_val
                if dept:
                    dept_liters[dept] += liters_val
            except (ValueError, TypeError):
                pass
        
        # Get amount
        amount = record.get('AMOUNT')
        if amount:
            try:
                amount_val = float(str(amount).replace(',', '').replace('KSH', '').strip())
                total_amount += amount_val
                car_amount[car] += amount_val
                if driver:
                    driver_amount[driver] += amount_val
                if dept:
                    dept_amount[dept] += amount_val
            except (ValueError, TypeError):
                pass
        
        # Get fuel type
        fuel_type = str(record.get('TYPE', '')).strip().upper()
        if fuel_type and liters:
            try:
                fuel_types[fuel_type] += float(str(liters).replace(',', ''))
            except:
                pass
        
        # Get odometer
        odometer = record.get('ODOMETER')
        datetime_str = record.get('DATETIME', '')
        if odometer:
            try:
                odo_val = int(float(str(odometer).replace(',', '')))
                car_odometer_readings[car].append((datetime_str, odo_val))
            except (ValueError, TypeError):
                pass
        
        # Daily breakdown
        if datetime_str:
            try:
                day_key = datetime_str[:10]  # YYYY-MM-DD
                if liters:
                    daily_liters[day_key] += float(str(liters).replace(',', ''))
                if amount:
                    daily_amount[day_key] += float(str(amount).replace(',', ''))
                daily_count[day_key] += 1
            except:
                pass
    
    # Find top consumer (most liters)
    top_consumer = None
    top_consumer_liters = 0
    for car, liters in car_liters.items():
        if liters > top_consumer_liters:
            top_consumer = car
            top_consumer_liters = liters
    
    # Find top driver (most liters)
    top_driver = None
    top_driver_liters = 0
    for driver, liters in driver_liters.items():
        if liters > top_driver_liters:
            top_driver = driver
            top_driver_liters = liters
    
    # Find top department (most liters)
    top_dept = None
    top_dept_liters = 0
    for dept, liters in dept_liters.items():
        if liters > top_dept_liters:
            top_dept = dept
            top_dept_liters = liters
    
    # Calculate distance traveled per car
    car_distance = {}
    for car, readings in car_odometer_readings.items():
        if len(readings) >= 2:
            readings_sorted = sorted(readings, key=lambda x: x[1])
            min_odo = readings_sorted[0][1]
            max_odo = readings_sorted[-1][1]
            distance = max_odo - min_odo
            if distance > 0:
                car_distance[car] = distance
    
    # Find car with most distance
    most_distance_car = None
    most_distance_km = 0
    for car, distance in car_distance.items():
        if distance > most_distance_km:
            most_distance_car = car
            most_distance_km = distance
    
    # Calculate efficiency (km per liter) for each car
    car_efficiency = {}
    for car in cars_fueled:
        if car in car_distance and car in car_liters:
            distance = car_distance[car]
            liters = car_liters[car]
            if liters > 0:
                efficiency = distance / liters
                car_efficiency[car] = efficiency
    
    # Find best efficiency (highest km/liter)
    best_efficiency_car = None
    best_efficiency_value = 0
    for car, efficiency in car_efficiency.items():
        if efficiency > best_efficiency_value:
            best_efficiency_car = car
            best_efficiency_value = efficiency
    
    # Find worst efficiency (lowest km/liter)
    worst_efficiency_car = None
    worst_efficiency_value = float('inf')
    for car, efficiency in car_efficiency.items():
        if efficiency < worst_efficiency_value:
            worst_efficiency_car = car
            worst_efficiency_value = efficiency
    
    # Find most frequently fueled car
    most_fueled_car = None
    most_fueled_count = 0
    for car, count in car_fuel_count.items():
        if count > most_fueled_count:
            most_fueled_car = car
            most_fueled_count = count
    
    # Calculate average price per liter
    avg_price_per_liter = total_amount / total_liters if total_liters > 0 else 0
    
    # Total distance covered by fleet
    total_distance = sum(car_distance.values())
    
    # Fleet average efficiency
    fleet_efficiency = total_distance / total_liters if total_liters > 0 else 0
    
    return {
        'period_days': days,
        'generated_at': datetime.now().isoformat(),
        'total_liters': round(total_liters, 2),
        'total_amount': round(total_amount, 2),
        'total_distance': total_distance,
        'cars_fueled_count': len(cars_fueled),
        'drivers_count': len(drivers),
        'departments_count': len(departments),
        'records_count': len(records),
        'avg_price_per_liter': round(avg_price_per_liter, 2),
        'fleet_efficiency': round(fleet_efficiency, 2),
        'fuel_types': dict(fuel_types),
        'daily_breakdown': {
            'liters': dict(daily_liters),
            'amount': dict(daily_amount),
            'count': dict(daily_count)
        },
        'top_consumer': {
            'car': top_consumer,
            'liters': round(top_consumer_liters, 2)
        } if top_consumer else None,
        'top_driver': {
            'name': top_driver,
            'liters': round(top_driver_liters, 2)
        } if top_driver else None,
        'top_department': {
            'name': top_dept,
            'liters': round(top_dept_liters, 2)
        } if top_dept else None,
        'most_distance': {
            'car': most_distance_car,
            'km': most_distance_km
        } if most_distance_car else None,
        'most_fueled': {
            'car': most_fueled_car,
            'count': most_fueled_count
        } if most_fueled_car else None,
        'best_efficiency': {
            'car': best_efficiency_car,
            'km_per_liter': round(best_efficiency_value, 2)
        } if best_efficiency_car else None,
        'worst_efficiency': {
            'car': worst_efficiency_car,
            'km_per_liter': round(worst_efficiency_value, 2)
        } if worst_efficiency_car and worst_efficiency_value != float('inf') else None,
        'dept_breakdown': {dept: round(liters, 2) for dept, liters in dept_liters.items()},
        'car_breakdown': {car: round(liters, 2) for car, liters in sorted(car_liters.items(), key=lambda x: -x[1])[:10]}
    }


def format_daily_summary(stats: Dict) -> str:
    """Format DAILY summary - focused on today's activity and quick insights."""
    if not stats:
        return "☀️ *DAILY FUEL REPORT*\n\n_No fuel records found for today._"
    
    now = datetime.now()
    
    msg = "☀️ *DAILY FUEL REPORT*\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += f"📅 {now.strftime('%A, %d %B %Y')}\n\n"
    
    # Today's totals
    msg += "💰 *TODAY'S TOTALS*\n"
    msg += f"   ⛽ Fuel: *{stats['total_liters']:,.2f} L*\n"
    msg += f"   💵 Spent: *KSH {stats['total_amount']:,.0f}*\n"
    msg += f"   📝 Reports: *{stats['records_count']}*\n"
    msg += f"   🚗 Vehicles: *{stats['cars_fueled_count']}*\n\n"
    
    # Price info
    if stats.get('avg_price_per_liter'):
        msg += f"📈 Avg Price: *KSH {stats['avg_price_per_liter']:.2f}/L*\n\n"
    
    # Fuel type breakdown
    if stats.get('fuel_types'):
        msg += "⛽ *BY FUEL TYPE*\n"
        for ftype, liters in sorted(stats['fuel_types'].items(), key=lambda x: -x[1]):
            msg += f"   • {ftype}: {liters:,.2f} L\n"
        msg += "\n"
    
    # Top vehicle today
    if stats.get('top_consumer'):
        tc = stats['top_consumer']
        msg += f"🏆 *TOP VEHICLE*\n"
        msg += f"   {tc['car']} ({tc['liters']:,.2f} L)\n\n"
    
    # Department activity
    if stats.get('dept_breakdown') and len(stats['dept_breakdown']) > 0:
        msg += "🏢 *DEPARTMENTS*\n"
        for dept, liters in sorted(stats['dept_breakdown'].items(), key=lambda x: -x[1]):
            msg += f"   • {dept}: {liters:,.2f} L\n"
        msg += "\n"
    
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += f"_Report generated at {now.strftime('%H:%M')}_"
    
    return msg


def format_weekly_summary(stats: Dict) -> str:
    """Format WEEKLY summary - comprehensive view with trends and rankings."""
    if not stats:
        return "📊 *WEEKLY FUEL SUMMARY*\n\n_No fuel records found for this week._"
    
    now = datetime.now()
    week_start = now - timedelta(days=7)
    
    msg = "📊 *WEEKLY FUEL SUMMARY*\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += f"📅 {week_start.strftime('%d %b')} - {now.strftime('%d %b %Y')}\n\n"
    
    # Week totals
    msg += "💰 *WEEK TOTALS*\n"
    msg += f"   ⛽ Total Fuel: *{stats['total_liters']:,.2f} L*\n"
    msg += f"   💵 Total Spent: *KSH {stats['total_amount']:,.0f}*\n"
    msg += f"   🛣️ Distance: *{stats.get('total_distance', 0):,} km*\n"
    msg += f"   📝 Reports: *{stats['records_count']}*\n\n"
    
    # Averages
    avg_daily_liters = stats['total_liters'] / 7
    avg_daily_spend = stats['total_amount'] / 7
    msg += "📈 *DAILY AVERAGES*\n"
    msg += f"   ⛽ {avg_daily_liters:,.1f} L/day\n"
    msg += f"   💵 KSH {avg_daily_spend:,.0f}/day\n\n"
    
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += "🏆 *TOP PERFORMERS*\n\n"
    
    # Top vehicle
    if stats.get('top_consumer'):
        tc = stats['top_consumer']
        msg += f"🚗 *Top Vehicle:*\n"
        msg += f"   {tc['car']} - {tc['liters']:,.2f} L\n\n"
    
    # Top driver
    if stats.get('top_driver'):
        td = stats['top_driver']
        msg += f"👤 *Top Driver:*\n"
        msg += f"   {td['name']} - {td['liters']:,.2f} L\n\n"
    
    # Top department
    if stats.get('top_department'):
        tdp = stats['top_department']
        msg += f"🏢 *Top Department:*\n"
        msg += f"   {tdp['name']} - {tdp['liters']:,.2f} L\n\n"
    
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += "📊 *EFFICIENCY STATS*\n\n"
    
    # Fleet efficiency
    if stats.get('fleet_efficiency'):
        msg += f"🚛 *Fleet Average:* {stats['fleet_efficiency']:.2f} km/L\n\n"
    
    # Best efficiency
    if stats.get('best_efficiency'):
        be = stats['best_efficiency']
        msg += f"✅ *Most Efficient:*\n"
        msg += f"   {be['car']} ({be['km_per_liter']:.2f} km/L)\n\n"
    
    # Worst efficiency
    if stats.get('worst_efficiency'):
        we = stats['worst_efficiency']
        msg += f"⚠️ *Least Efficient:*\n"
        msg += f"   {we['car']} ({we['km_per_liter']:.2f} km/L)\n\n"
    
    # Most distance
    if stats.get('most_distance'):
        md = stats['most_distance']
        msg += f"🛣️ *Most Distance:*\n"
        msg += f"   {md['car']} ({md['km']:,} km)\n"
    
    msg += "\n━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += "_Weekly auto-report_"
    
    return msg


def format_monthly_summary(stats: Dict) -> str:
    """Format MONTHLY summary - executive overview with department breakdown and trends."""
    if not stats:
        return "📈 *MONTHLY FUEL REPORT*\n\n_No fuel records found for this month._"
    
    now = datetime.now()
    month_start = now - timedelta(days=30)
    
    msg = "📈 *MONTHLY FUEL REPORT*\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += f"📅 {month_start.strftime('%d %b')} - {now.strftime('%d %b %Y')}\n\n"
    
    # Executive Summary
    msg += "💼 *EXECUTIVE SUMMARY*\n"
    msg += "─────────────────────────\n"
    msg += f"⛽ Total Fuel: *{stats['total_liters']:,.2f} L*\n"
    msg += f"💵 Total Cost: *KSH {stats['total_amount']:,.0f}*\n"
    msg += f"🛣️ Total Distance: *{stats.get('total_distance', 0):,} km*\n"
    msg += f"🚛 Fleet Efficiency: *{stats.get('fleet_efficiency', 0):.2f} km/L*\n"
    msg += f"📈 Avg Price: *KSH {stats.get('avg_price_per_liter', 0):.2f}/L*\n\n"
    
    # Monthly averages
    avg_weekly = stats['total_liters'] / 4
    avg_daily = stats['total_liters'] / 30
    msg += "📊 *CONSUMPTION RATES*\n"
    msg += "─────────────────────────\n"
    msg += f"📅 Daily: {avg_daily:,.1f} L/day\n"
    msg += f"📆 Weekly: {avg_weekly:,.1f} L/week\n"
    msg += f"💰 Daily Cost: KSH {stats['total_amount']/30:,.0f}\n\n"
    
    # Fleet stats
    msg += "🚗 *FLEET OVERVIEW*\n"
    msg += "─────────────────────────\n"
    msg += f"Vehicles Active: *{stats['cars_fueled_count']}*\n"
    msg += f"Drivers Active: *{stats.get('drivers_count', 0)}*\n"
    msg += f"Departments: *{stats.get('departments_count', 0)}*\n"
    msg += f"Total Reports: *{stats['records_count']}*\n\n"
    
    # Fuel type distribution
    if stats.get('fuel_types'):
        msg += "⛽ *FUEL DISTRIBUTION*\n"
        msg += "─────────────────────────\n"
        for ftype, liters in sorted(stats['fuel_types'].items(), key=lambda x: -x[1]):
            pct = (liters / stats['total_liters'] * 100) if stats['total_liters'] > 0 else 0
            msg += f"• {ftype}: {liters:,.0f} L ({pct:.1f}%)\n"
        msg += "\n"
    
    # Department breakdown
    if stats.get('dept_breakdown') and len(stats['dept_breakdown']) > 0:
        msg += "🏢 *BY DEPARTMENT*\n"
        msg += "─────────────────────────\n"
        for dept, liters in sorted(stats['dept_breakdown'].items(), key=lambda x: -x[1]):
            pct = (liters / stats['total_liters'] * 100) if stats['total_liters'] > 0 else 0
            msg += f"• {dept}: {liters:,.0f} L ({pct:.1f}%)\n"
        msg += "\n"
    
    # Top 5 vehicles
    if stats.get('car_breakdown'):
        msg += "🔝 *TOP 5 VEHICLES*\n"
        msg += "─────────────────────────\n"
        for i, (car, liters) in enumerate(list(stats['car_breakdown'].items())[:5], 1):
            msg += f"{i}. {car}: {liters:,.2f} L\n"
        msg += "\n"
    
    msg += "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    
    # Performance highlights
    msg += "🌟 *HIGHLIGHTS*\n"
    msg += "─────────────────────────\n"
    
    if stats.get('top_consumer'):
        msg += f"🏆 Top Consumer: {stats['top_consumer']['car']}\n"
    
    if stats.get('most_distance'):
        msg += f"🛣️ Most Active: {stats['most_distance']['car']}\n"
    
    if stats.get('best_efficiency'):
        msg += f"✅ Most Efficient: {stats['best_efficiency']['car']}\n"
    
    if stats.get('worst_efficiency'):
        msg += f"⚠️ Needs Review: {stats['worst_efficiency']['car']}\n"
    
    if stats.get('most_fueled'):
        mf = stats['most_fueled']
        msg += f"🔄 Most Refueled: {mf['car']} ({mf['count']}x)\n"
    
    msg += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    msg += f"_Monthly report • {now.strftime('%B %Y')}_"
    
    return msg


def format_summary_message(stats: Dict, days: int = 7) -> str:
    """Format statistics based on the period."""
    if days <= 1:
        return format_daily_summary(stats)
    elif days <= 7:
        return format_weekly_summary(stats)
    else:
        return format_monthly_summary(stats)


def save_summary_for_notification(stats: Dict, message: str):
    """Save summary to JSON file for Node.js listener to send."""
    summary_data = {
        'timestamp': datetime.now().isoformat(),
        'stats': stats,
        'message': message,
        'sent': False
    }
    
    # Ensure directory exists
    SUMMARY_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    
    with open(SUMMARY_OUTPUT_PATH, 'w') as f:
        json.dump(summary_data, f, indent=2)
    
    logger.info(f"Summary saved to {SUMMARY_OUTPUT_PATH}")


def generate_weekly_summary(days: int = 7, save_for_whatsapp: bool = True) -> Tuple[Dict, str]:
    """Generate summary for the specified period and optionally save for WhatsApp notification."""
    period_name = "daily" if days <= 1 else ("weekly" if days <= 7 else "monthly")
    logger.info(f"Generating {period_name} fuel summary for the last {days} days...")
    
    records = load_weekly_data(days)
    stats = calculate_statistics(records, days)
    message = format_summary_message(stats, days)
    
    if save_for_whatsapp:
        save_summary_for_notification(stats, message)
    
    return stats, message


def get_car_summary(plate: str, days: int = 30) -> Tuple[Dict, str]:
    """
    Get summary for a specific vehicle.
    Returns statistics and formatted message for the car.
    """
    import re
    
    # Normalize plate
    normalized_plate = re.sub(r'\s+', '', str(plate)).upper()
    
    excel_path = get_excel_path()
    if not excel_path or not excel_path.exists():
        return None, f"❌ Excel file not found"
    
    try:
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        
        records = []
        cutoff_date = datetime.now() - timedelta(days=days)
        
        # Get column headers from first row
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            record = dict(zip(headers, row))
            
            # Check if this record is for our car
            row_car = str(record.get('CAR', '')).upper().replace(' ', '')
            if row_car != normalized_plate:
                continue
            
            # Parse datetime to check if within range
            datetime_str = record.get('DATETIME', '')
            if datetime_str:
                try:
                    record_date = datetime.strptime(datetime_str, '%Y-%m-%d-%H-%M')
                    if record_date >= cutoff_date:
                        records.append(record)
                except ValueError:
                    records.append(record)  # Include if can't parse date
        
        wb.close()
        
        if not records:
            return None, f"📋 No records found for *{normalized_plate}* in the last {days} days."
        
        # Calculate car-specific stats
        total_liters = 0.0
        total_amount = 0.0
        fuel_types = defaultdict(float)
        drivers = set()
        departments = set()
        odometer_readings = []
        
        for rec in records:
            try:
                liters = float(str(rec.get('LITERS', 0)).replace(',', ''))
                total_liters += liters
                fuel_type = rec.get('TYPE', 'UNKNOWN')
                fuel_types[fuel_type] += liters
            except:
                pass
            
            try:
                amount = float(str(rec.get('AMOUNT', 0)).replace(',', ''))
                total_amount += amount
            except:
                pass
            
            if rec.get('DRIVER'):
                drivers.add(rec['DRIVER'])
            if rec.get('DEPARTMENT'):
                departments.add(rec['DEPARTMENT'])
            
            try:
                odo = int(float(str(rec.get('ODOMETER', 0)).replace(',', '')))
                if odo > 0:
                    odometer_readings.append(odo)
            except:
                pass
        
        # Calculate distance traveled
        distance = 0
        if len(odometer_readings) >= 2:
            distance = max(odometer_readings) - min(odometer_readings)
        
        # Calculate efficiency (km/L)
        efficiency = distance / total_liters if total_liters > 0 and distance > 0 else 0
        
        # Build stats dict
        stats = {
            'plate': normalized_plate,
            'records': len(records),
            'total_liters': total_liters,
            'total_amount': total_amount,
            'fuel_types': dict(fuel_types),
            'drivers': list(drivers),
            'departments': list(departments),
            'min_odometer': min(odometer_readings) if odometer_readings else 0,
            'max_odometer': max(odometer_readings) if odometer_readings else 0,
            'distance': distance,
            'efficiency': efficiency,
            'days': days
        }
        
        # Format message
        msg = f"🚗 *VEHICLE SUMMARY: {normalized_plate}*\n"
        msg += f"━━━━━━━━━━━━━━━━━━━━━━━\n"
        msg += f"_Last {days} days_\n\n"
        
        msg += f"📊 *OVERVIEW*\n"
        msg += f"• Fuel Records: {len(records)}\n"
        msg += f"• Total Fuel: {total_liters:,.1f} L\n"
        msg += f"• Total Spent: KSH {total_amount:,.0f}\n"
        
        if distance > 0:
            msg += f"• Distance: {distance:,} km\n"
            msg += f"• Efficiency: {efficiency:.1f} km/L\n"
        
        msg += f"\n⛽ *FUEL BREAKDOWN*\n"
        for fuel_type, liters in sorted(fuel_types.items(), key=lambda x: -x[1]):
            pct = (liters / total_liters * 100) if total_liters > 0 else 0
            msg += f"• {fuel_type}: {liters:,.1f} L ({pct:.0f}%)\n"
        
        if odometer_readings:
            msg += f"\n📏 *ODOMETER*\n"
            msg += f"• First: {min(odometer_readings):,} km\n"
            msg += f"• Latest: {max(odometer_readings):,} km\n"
        
        if drivers:
            msg += f"\n👤 *DRIVER(S)*\n"
            for driver in sorted(drivers):
                msg += f"• {driver}\n"
        
        if departments:
            msg += f"\n🏢 *DEPARTMENT(S)*\n"
            for dept in sorted(departments):
                msg += f"• {dept}\n"
        
        msg += f"\n━━━━━━━━━━━━━━━━━━━━━━━"
        
        return stats, msg
        
    except Exception as e:
        logger.error(f"Error getting car summary: {e}")
        return None, f"❌ Error getting summary: {e}"


CAR_SUMMARY_OUTPUT_PATH = ROOT_DIR / 'data' / 'car_summary.json'


def save_car_summary_for_notification(plate: str, stats: Dict, message: str):
    """Save car summary for WhatsApp notification."""
    summary_data = {
        'plate': plate,
        'generated_at': datetime.now().isoformat(),
        'stats': stats,
        'message': message,
        'sent': False
    }
    
    CAR_SUMMARY_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    
    with open(CAR_SUMMARY_OUTPUT_PATH, 'w') as f:
        json.dump(summary_data, f, indent=2)
    
    logger.info(f"Car summary saved to {CAR_SUMMARY_OUTPUT_PATH}")


if __name__ == '__main__':
    import sys
    
    print("\n" + "=" * 60)
    print("  WhatsApp Fuel Extractor - Summary Generator")
    print("=" * 60 + "\n")
    
    # Check if this is a car summary request
    if len(sys.argv) > 1 and sys.argv[1] == '--car':
        if len(sys.argv) < 3:
            print("Usage: python weekly_summary.py --car <PLATE> [days]")
            sys.exit(1)
        
        plate = sys.argv[2]
        days = 30
        if len(sys.argv) > 3:
            try:
                days = int(sys.argv[3])
            except ValueError:
                pass
        
        stats, message = get_car_summary(plate, days)
        if stats:
            save_car_summary_for_notification(plate, stats, message)
        print("\n" + message)
        print("\n" + "=" * 60 + "\n")
        sys.exit(0)
    
    # Default: period summary
    days = 7
    if len(sys.argv) > 1:
        try:
            days = int(sys.argv[1])
        except ValueError:
            print(f"Usage: {sys.argv[0]} [days] OR {sys.argv[0]} --car <PLATE> [days]")
            sys.exit(1)
    
    stats, message = generate_weekly_summary(days)
    
    print("\n" + message)
    print("\n" + "=" * 60)
    print("Summary saved for WhatsApp notification.")
    print("=" * 60 + "\n")
