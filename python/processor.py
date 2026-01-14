#!/usr/bin/env python3
"""
WhatsApp Fuel Extractor - Message Processor

Reads raw message files from the Node.js listener, extracts fuel data,
validates fields, and appends records to an Excel file.

Features:
- Robust regex parsing handles many format variations
- Case-insensitive field matching
- Handles fields in any order
- Validates data types and ranges
- Odometer validation (must be > previous reading for same car)
- Error logging for failed parses
- Sends validation errors back to WhatsApp group
- Hourly scheduling support
"""

import os
import sys
import re
import json
import shutil
import logging
import fcntl
from contextlib import contextmanager
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# Thread-safe file operations
@contextmanager
def file_lock(filepath: Path, mode: str = 'r'):
    """Context manager for thread-safe file access with locking"""
    f = None
    try:
        f = open(filepath, mode)
        fcntl.flock(f.fileno(), fcntl.LOCK_EX if 'w' in mode else fcntl.LOCK_SH)
        yield f
    finally:
        if f:
            fcntl.flock(f.fileno(), fcntl.LOCK_UN)
            f.close()


def safe_json_load(filepath: Path, default: Any = None) -> Any:
    """Safely load JSON with error handling and locking"""
    if not filepath.exists():
        return default if default is not None else {}
    try:
        with file_lock(filepath, 'r') as f:
            content = f.read()
            if not content.strip():
                return default if default is not None else {}
            return json.loads(content)
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in {filepath}: {e}")
        # Create backup of corrupted file
        backup_path = filepath.with_suffix('.json.corrupted')
        try:
            shutil.copy2(filepath, backup_path)
            logger.info(f"Created backup at {backup_path}")
        except:
            pass
        return default if default is not None else {}
    except Exception as e:
        logger.error(f"Error loading {filepath}: {e}")
        return default if default is not None else {}


def safe_json_save(filepath: Path, data: Any, indent: int = 2) -> bool:
    """Safely save JSON with atomic write and locking"""
    temp_path = filepath.with_suffix('.json.tmp')
    try:
        # Ensure parent directory exists
        filepath.parent.mkdir(parents=True, exist_ok=True)
        # Write to temp file first (atomic write pattern)
        with open(temp_path, 'w') as f:
            json.dump(data, f, indent=indent, default=str)
        # Rename temp to actual (atomic on most filesystems)
        temp_path.rename(filepath)
        return True
    except Exception as e:
        logger.error(f"Error saving {filepath}: {e}")
        # Clean up temp file if exists
        if temp_path.exists():
            try:
                temp_path.unlink()
            except:
                pass
        return False

# Setup logging with rotation
# Rotate log files: max 5MB per file, keep 5 backup files
LOG_FILE = Path(__file__).parent.parent / 'processor.log'
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# Create handlers
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)

file_handler = RotatingFileHandler(
    LOG_FILE,
    maxBytes=5 * 1024 * 1024,  # 5 MB
    backupCount=5,
    encoding='utf-8'
)
file_handler.setFormatter(log_formatter)

# Configure logger
logging.basicConfig(level=logging.INFO, handlers=[console_handler, file_handler])
logger = logging.getLogger(__name__)

# Paths
ROOT_DIR = Path(__file__).parent.parent
# Ensure local python/ modules are importable
if str(Path(__file__).parent) not in sys.path:
    sys.path.append(str(Path(__file__).parent))
try:
    from env import load_env
    load_env()
except Exception:
    pass
CONFIG_PATH = ROOT_DIR / 'config.json'
RAW_MESSAGES_PATH = ROOT_DIR / 'data' / 'raw_messages'
PROCESSED_PATH = ROOT_DIR / 'data' / 'processed'
ERRORS_PATH = ROOT_DIR / 'data' / 'errors'
VALIDATION_ERRORS_PATH = ROOT_DIR / 'data' / 'validation_errors.json'
PENDING_APPROVALS_PATH = ROOT_DIR / 'data' / 'pending_approvals.json'
CAR_LAST_UPDATE_PATH = ROOT_DIR / 'data' / 'car_last_update.json'

# Timeouts for validation (in hours)
CAR_COOLDOWN_HOURS = 12  # Same car can't fuel within 12 hours
EDIT_APPROVAL_TIMEOUT_MINUTES = 10

# Fuel efficiency thresholds (km per liter)
EFFICIENCY_ALERT_LOW = 4.0   # Below this = alert (possible fuel theft or vehicle issue)
EFFICIENCY_ALERT_HIGH = 20.0  # Above this = suspicious (possible odometer tampering)
EFFICIENCY_GOOD_MIN = 6.0     # Minimum for "good" efficiency
EFFICIENCY_GOOD_MAX = 12.0    # Maximum for "good" efficiency

# Efficiency tracking file
EFFICIENCY_HISTORY_PATH = ROOT_DIR / 'data' / 'efficiency_history.json'

# Allowed vehicle registration numbers (normalized: uppercase, no spaces)
ALLOWED_PLATES = {
    'KCA542Q', 'KCB711C', 'KCE090R', 'KCE690F', 'KCE699F', 'KCG668W', 'KCH167M',
    'KCQ215F', 'KCQ581M', 'KCQ618K', 'KCU938R', 'KCU729C', 'KCY076X', 'KCY080X',
    'KCY084X', 'KCY090X', 'KCY838X', 'KCZ154S', 'KCZ155P', 'KCZ181P', 'KCZ199P',
    'KCZ223P', 'KCZ476E', 'KCZ751V', 'KDA609E', 'KDA717B', 'KDB323M', 'KDB585E',
    'KDC207R', 'KDC490Q', 'KDC739F', 'KDD684Y', 'KDD689Y', 'KDE264M', 'KDE638J',
    'KDK728K', 'KDK732K', 'KDK780K', 'KDK815R', 'KDM306S', 'KDM308S', 'KDM309S',
    'KDM794R', 'KDM840V', 'KDR592N', 'KDR594N', 'KDS453Y', 'KDS525D', 'KDS919Y',
    'KDT728R', 'KDT916R', 'KDT923R', 'KMDG902Z', 'KMEL225X', 'KMFF099Z', 'KMFF113Z',
    'KMFF162Z', 'KMGK596V', 'KMGS239H', 'KCG669W', 'KDP655F', 'KDS949Y', 'KDT724R',
    'KCK201X', 'KCK686A', 'KCL502T', 'KCN496A', 'KCU237Z', 'KCY930Y', 'KDD655F',
    'KDN753G', 'KDN759G', 'KDU613B', 'UA234BJ', 'KDT794R', 'KCP337X', 'KDM402L',
    'KDV064S', 'KDV072L', 'KDV438W', 'KDV439W', 'KDV437W'
}


def save_validation_error(car: str, driver: str, issue: str, sender_phone: str = '', is_approval_request: bool = False):
    """
    Save a validation error for the Node.js listener to send to WhatsApp.
    
    Args:
        car: Vehicle plate
        driver: Driver name
        issue: Error message to send
        sender_phone: Phone number of the sender (for tagging)
        is_approval_request: If True, tag admins instead of sender
    """
    error = {
        'timestamp': datetime.now().isoformat(),
        'car': car,
        'driver': driver,
        'issue': issue,
        'sender_phone': sender_phone,
        'is_approval_request': is_approval_request,
        'notified': False
    }
    
    errors = safe_json_load(VALIDATION_ERRORS_PATH, [])
    if not isinstance(errors, list):
        errors = []
    
    errors.append(error)
    
    # Limit to last 1000 errors to prevent file bloat
    if len(errors) > 1000:
        errors = errors[-1000:]
    
    safe_json_save(VALIDATION_ERRORS_PATH, errors)
    logger.info(f"[SAVED] Saved validation error for notification: {car} - {issue[:50]}...")


def save_pending_approval(approval_type: str, record: Dict, original_record: Optional[Dict] = None, reason: str = ''):
    """Save a record that needs admin approval before processing."""
    import uuid
    
    approval = {
        'id': str(uuid.uuid4())[:8],
        'type': approval_type,  # 'edit', 'driver_change', 'duplicate_check'
        'timestamp': datetime.now().isoformat(),
        'record': record,
        'original_record': original_record,
        'reason': reason,
        'status': 'pending',
        'notified': False
    }
    
    approvals = safe_json_load(PENDING_APPROVALS_PATH, [])
    if not isinstance(approvals, list):
        approvals = []
    
    approvals.append(approval)
    
    # Clean up old approvals (keep last 500)
    if len(approvals) > 500:
        # Keep only recent and pending approvals
        pending = [a for a in approvals if a.get('status') == 'pending']
        processed = [a for a in approvals if a.get('status') != 'pending'][-300:]
        approvals = pending + processed
    
    safe_json_save(PENDING_APPROVALS_PATH, approvals)
    logger.info(f"[SAVED] Saved pending approval ({approval_type}): {approval['id']} - {reason}")
    return approval['id']


def get_pending_approvals() -> List[Dict]:
    """Get all pending approvals."""
    approvals = safe_json_load(PENDING_APPROVALS_PATH, [])
    if not isinstance(approvals, list):
        return []
    return [a for a in approvals if isinstance(a, dict) and a.get('status') == 'pending']


def approve_pending(approval_id: str) -> Tuple[bool, str, Optional[Dict]]:
    """Approve a pending record. Returns (success, message, record)."""
    approvals = safe_json_load(PENDING_APPROVALS_PATH, [])
    if not approvals:
        return False, 'No pending approvals', None
    
    for approval in approvals:
        if approval.get('id') == approval_id:
            if approval.get('status') != 'pending':
                return False, f'Approval {approval_id} already processed', None
            
            approval['status'] = 'approved'
            approval['approved_at'] = datetime.now().isoformat()
            
            if safe_json_save(PENDING_APPROVALS_PATH, approvals):
                return True, f'Approval {approval_id} approved', approval.get('record')
            else:
                return False, 'Failed to save approval', None
    
    return False, f'Approval {approval_id} not found', None


def reject_pending(approval_id: str) -> Tuple[bool, str]:
    """Reject a pending record."""
    approvals = safe_json_load(PENDING_APPROVALS_PATH, [])
    if not approvals:
        return False, 'No pending approvals'
    
    for approval in approvals:
        if approval.get('id') == approval_id:
            if approval.get('status') != 'pending':
                return False, f'Approval {approval_id} already processed'
            
            approval['status'] = 'rejected'
            approval['rejected_at'] = datetime.now().isoformat()
            
            if safe_json_save(PENDING_APPROVALS_PATH, approvals):
                return True, f'Approval {approval_id} rejected'
            else:
                return False, 'Failed to save rejection'
    
    return False, f'Approval {approval_id} not found'


def get_car_last_update(car_plate: str) -> Optional[Dict]:
    """Get the last update info for a car plate."""
    updates = safe_json_load(CAR_LAST_UPDATE_PATH, {})
    if not isinstance(updates, dict):
        return None
    return updates.get(normalize_plate(car_plate))


def update_car_last_update(car_plate: str, record: Dict, efficiency: Optional[float] = None):
    """Update the last fuel record timestamp for a car plate."""
    updates = safe_json_load(CAR_LAST_UPDATE_PATH, {})
    if not isinstance(updates, dict):
        updates = {}
    
    updates[normalize_plate(car_plate)] = {
        'timestamp': datetime.now().isoformat(),
        'driver': record.get('driver', ''),
        'liters': record.get('liters', ''),
        'amount': record.get('amount', ''),
        'odometer': record.get('odometer', ''),
        'type': record.get('type', ''),
        'department': record.get('department', ''),
        'efficiency': efficiency  # km/L for this fill-up
    }
    
    safe_json_save(CAR_LAST_UPDATE_PATH, updates)


def calculate_fuel_efficiency(car_plate: str, current_odometer: int, current_liters: float) -> Tuple[Optional[float], Optional[Dict]]:
    """
    Calculate fuel efficiency (km/L) based on previous and current odometer readings.
    Returns (efficiency_km_per_liter, alert_info_if_needed)
    """
    last_update = get_car_last_update(car_plate)
    
    if not last_update:
        return None, None  # First record, can't calculate
    
    try:
        # Get previous odometer
        prev_odo = int(float(str(last_update.get('odometer', 0)).replace(',', '')))
        prev_liters = float(str(last_update.get('liters', 0)).replace(',', ''))
        
        if prev_odo <= 0 or current_odometer <= prev_odo:
            return None, None  # Can't calculate without valid odometer progression
        
        # Distance traveled since last fuel-up
        distance = current_odometer - prev_odo
        
        # Efficiency = distance traveled / fuel used in PREVIOUS fill-up
        # (This fill-up will be used for the NEXT calculation)
        if prev_liters <= 0:
            return None, None
        
        efficiency = distance / prev_liters
        
        # Check for alerts
        alert = None
        if efficiency < EFFICIENCY_ALERT_LOW:
            alert = {
                'type': 'low_efficiency',
                'severity': 'warning',
                'efficiency': efficiency,
                'distance': distance,
                'liters': prev_liters,
                'message': f'Low fuel efficiency: {efficiency:.1f} km/L (expected: {EFFICIENCY_GOOD_MIN}-{EFFICIENCY_GOOD_MAX} km/L). Possible fuel theft or vehicle issue.'
            }
        elif efficiency > EFFICIENCY_ALERT_HIGH:
            alert = {
                'type': 'high_efficiency',
                'severity': 'warning',
                'efficiency': efficiency,
                'distance': distance,
                'liters': prev_liters,
                'message': f'Unusually high efficiency: {efficiency:.1f} km/L. Possible odometer discrepancy.'
            }
        
        return efficiency, alert
        
    except (ValueError, TypeError) as e:
        logger.error(f"Error calculating efficiency for {car_plate}: {e}")
        return None, None


def save_efficiency_record(car_plate: str, efficiency: float, distance: int, liters: float, driver: str):
    """Save efficiency record for historical tracking."""
    history = safe_json_load(EFFICIENCY_HISTORY_PATH, [])
    if not isinstance(history, list):
        history = []
    
    record = {
        'timestamp': datetime.now().isoformat(),
        'car': normalize_plate(car_plate),
        'driver': driver,
        'efficiency': round(efficiency, 2),
        'distance': distance,
        'liters': round(liters, 2)
    }
    
    history.append(record)
    
    # Keep last 5000 records
    if len(history) > 5000:
        history = history[-5000:]
    
    safe_json_save(EFFICIENCY_HISTORY_PATH, history)


def get_vehicle_efficiency_stats(car_plate: str, days: int = 30) -> Dict:
    """Get efficiency statistics for a vehicle."""
    history = safe_json_load(EFFICIENCY_HISTORY_PATH, [])
    if not isinstance(history, list):
        return {}
    
    cutoff = datetime.now() - timedelta(days=days)
    normalized_plate = normalize_plate(car_plate)
    
    vehicle_records = []
    for record in history:
        if record.get('car') != normalized_plate:
            continue
        try:
            record_time = datetime.fromisoformat(record.get('timestamp', ''))
            if record_time >= cutoff:
                vehicle_records.append(record)
        except:
            pass
    
    if not vehicle_records:
        return {'car': normalized_plate, 'records': 0}
    
    efficiencies = [r['efficiency'] for r in vehicle_records if r.get('efficiency')]
    total_distance = sum(r.get('distance', 0) for r in vehicle_records)
    total_liters = sum(r.get('liters', 0) for r in vehicle_records)
    
    return {
        'car': normalized_plate,
        'records': len(vehicle_records),
        'avg_efficiency': round(sum(efficiencies) / len(efficiencies), 2) if efficiencies else 0,
        'min_efficiency': round(min(efficiencies), 2) if efficiencies else 0,
        'max_efficiency': round(max(efficiencies), 2) if efficiencies else 0,
        'total_distance': total_distance,
        'total_liters': round(total_liters, 2),
        'days': days
    }


def save_efficiency_alert(car_plate: str, driver: str, alert: Dict):
    """Save efficiency alert for admin notification via WhatsApp."""
    msg = f"[!] *FUEL EFFICIENCY ALERT - {car_plate}*\n"
    msg += f"----------------------------\n\n"
    msg += f"Driver: {driver}\n"
    msg += f"Efficiency: *{alert['efficiency']:.1f} km/L*\n"
    msg += f"Distance: {alert['distance']:,} km\n"
    msg += f"Fuel Used: {alert['liters']:.1f} L\n\n"
    
    if alert['type'] == 'low_efficiency':
        msg += f"[!] *LOW EFFICIENCY WARNING*\n"
        msg += f"Expected range: {EFFICIENCY_GOOD_MIN}-{EFFICIENCY_GOOD_MAX} km/L\n"
        msg += f"Possible causes:\n"
        msg += f"- Fuel siphoning/theft\n"
        msg += f"- Vehicle mechanical issues\n"
        msg += f"- Incorrect odometer reading\n"
    else:
        msg += f"[?] *UNUSUALLY HIGH EFFICIENCY*\n"
        msg += f"This may indicate:\n"
        msg += f"- Odometer tampering\n"
        msg += f"- Data entry error\n"
    
    msg += f"\n_Please investigate_"
    
    save_validation_error(car_plate, driver, msg, '', is_approval_request=True)


def check_car_cooldown(car_plate: str, record: Dict) -> Tuple[bool, Optional[str]]:
    """
    Check if the same car is trying to fuel within 12 hours.
    Returns (can_proceed, approval_id_if_needed)
    """
    last_update = get_car_last_update(car_plate)
    
    if not last_update:
        # First record for this car - no cooldown
        return True, None
    
    last_timestamp = last_update.get('timestamp', '')
    
    try:
        last_time = datetime.fromisoformat(last_timestamp)
        hours_since = (datetime.now() - last_time).total_seconds() / 3600
        minutes_since = (datetime.now() - last_time).total_seconds() / 60
        
        if hours_since < CAR_COOLDOWN_HOURS:
            hours_remaining = CAR_COOLDOWN_HOURS - hours_since
            
            # Get previous and current values for comparison
            last_driver = last_update.get('driver', 'Unknown')
            new_driver = record.get('driver', 'Unknown')
            
            # Parse odometer values
            try:
                last_odo = int(float(str(last_update.get('odometer', 0)).replace(',', '')))
            except:
                last_odo = 0
            try:
                new_odo = int(float(str(record.get('odometer', 0)).replace(',', '')))
            except:
                new_odo = 0
            
            distance_traveled = new_odo - last_odo if new_odo > last_odo else 0
            
            # Parse liter values
            try:
                last_liters = float(str(last_update.get('liters', 0)).replace(',', ''))
            except:
                last_liters = 0
            try:
                new_liters = float(str(record.get('liters', 0)).replace(',', ''))
            except:
                new_liters = 0
            
            # Parse amount values
            try:
                last_amount = float(str(last_update.get('amount', 0)).replace(',', ''))
            except:
                last_amount = 0
            try:
                new_amount = float(str(record.get('amount', 0)).replace(',', ''))
            except:
                new_amount = 0
            
            # Format time interval nicely
            if hours_since < 1:
                time_interval = f"{int(minutes_since)} minutes"
            else:
                time_interval = f"{hours_since:.1f} hours"
            
            reason = f"Same car {car_plate} fueled {time_interval} ago (cooldown: {CAR_COOLDOWN_HOURS}h)"
            
            approval_id = save_pending_approval(
                'car_cooldown',
                record,
                last_update,
                reason
            )
            
            # Build detailed notification message
            msg = f"[!] *DUPLICATE FUEL REPORT - {car_plate}*\n"
            msg += f"----------------------------\n\n"
            
            msg += f"[TIME] *TIME SINCE LAST FUELING:* {time_interval}\n"
            msg += f"[WAIT] Cooldown remaining: {hours_remaining:.1f} hours\n\n"
            
            msg += f"[DRIVER] *DRIVER COMPARISON*\n"
            msg += f"- Previous: {last_driver}\n"
            msg += f"- Current: {new_driver}\n"
            if last_driver.lower().strip() != new_driver.lower().strip():
                msg += f"[!] _Driver changed!_\n"
            msg += "\n"
            
            msg += f"[ODO] *ODOMETER / DISTANCE*\n"
            msg += f"- Previous: {last_odo:,} km\n"
            msg += f"- Current: {new_odo:,} km\n"
            if distance_traveled > 0:
                msg += f"- Distance traveled: *{distance_traveled:,} km*\n"
            elif new_odo <= last_odo and new_odo > 0:
                msg += f"[!] _Odometer hasn't increased!_\n"
            msg += "\n"
            
            msg += f"[FUEL] *FUEL COMPARISON*\n"
            msg += f"- Previous: {last_liters:.1f} L (KSH {last_amount:,.0f})\n"
            msg += f"- Current: {new_liters:.1f} L (KSH {new_amount:,.0f})\n"
            
            # Calculate efficiency if we have distance
            if distance_traveled > 0 and last_liters > 0:
                efficiency = distance_traveled / last_liters
                msg += f"- Efficiency since last: {efficiency:.1f} km/L\n"
            msg += "\n"
            
            msg += f"----------------------------\n"
            msg += f"[ID] Approval ID: *{approval_id}*\n\n"
            msg += f"[OK] *!approve {approval_id}* - Log as new record\n"
            msg += f"[X] *!reject {approval_id}* - Discard"
            
            # Save as approval request - this will tag admins instead of the sender
            save_validation_error(car_plate, record.get('driver', ''), msg, record.get('sender_phone', ''), is_approval_request=True)
            
            return False, approval_id
            
    except Exception as e:
        logger.error(f"Error checking car cooldown: {e}")
    
    return True, None


# Normalize plate by removing spaces and uppercasing
def normalize_plate(plate: str) -> str:
    return re.sub(r'\s+', '', str(plate)).upper()


def is_allowed_plate(plate: str) -> bool:
    return normalize_plate(plate) in ALLOWED_PLATES


# Path for confirmation messages
CONFIRMATIONS_PATH = ROOT_DIR / 'data' / 'confirmations.json'


def save_confirmation(record: Dict, sender: str, efficiency: Optional[float] = None, distance: Optional[int] = None):
    """Save a confirmation message for the Node.js listener to send to WhatsApp."""
    # Format datetime for display
    dt_str = record.get('datetime', '')
    
    # Build clean confirmation message
    msg = "[LOGGED] *FUEL REPORT LOGGED*\n\n"
    
    if record.get('department'):
        msg += f"Department: {record['department']}\n"
    if record.get('driver'):
        msg += f"Driver: {record['driver']}\n"
    if record.get('car'):
        msg += f"Vehicle: {record['car']}\n"
    if record.get('liters'):
        try:
            liters = float(str(record['liters']).replace(',', ''))
            msg += f"Fuel: {liters:.2f} L"
        except:
            msg += f"Fuel: {record['liters']} L"
        if record.get('type'):
            msg += f" ({record['type']})"
        msg += "\n"
    if record.get('amount'):
        try:
            amount = float(str(record['amount']).replace(',', ''))
            msg += f"Amount: KSH {amount:,.0f}\n"
        except:
            msg += f"Amount: KSH {record['amount']}\n"
    if record.get('odometer'):
        try:
            odo = int(float(str(record['odometer']).replace(',', '')))
            msg += f"Odometer: {odo:,} km\n"
        except:
            msg += f"Odometer: {record['odometer']} km\n"
    
    # Add fuel efficiency if calculated
    if efficiency is not None and distance is not None:
        msg += f"\n[STATS] *Fuel Efficiency*\n"
        msg += f"Distance since last fill: {distance:,} km\n"
        msg += f"Efficiency: *{efficiency:.1f} km/L*\n"
        # Add efficiency rating
        if EFFICIENCY_GOOD_MIN <= efficiency <= EFFICIENCY_GOOD_MAX:
            msg += f"Rating: Good\n"
        elif efficiency < EFFICIENCY_ALERT_LOW:
            msg += f"Rating: Poor (check vehicle)\n"
        elif efficiency > EFFICIENCY_ALERT_HIGH:
            msg += f"Rating: Unusually high\n"
        else:
            msg += f"Rating: Normal\n"
    
    msg += f"\n_{dt_str} | {sender}_"
    
    confirmation = {
        'timestamp': datetime.now().isoformat(),
        'message': msg,
        'notified': False
    }
    
    confirmations = safe_json_load(CONFIRMATIONS_PATH, [])
    if not isinstance(confirmations, list):
        confirmations = []
    
    confirmations.append(confirmation)
    
    # Limit to last 500 confirmations to prevent file bloat
    if len(confirmations) > 500:
        confirmations = confirmations[-500:]
    
    safe_json_save(CONFIRMATIONS_PATH, confirmations)
    logger.info(f"[SAVED] Saved confirmation for: {record.get('car', 'N/A')}")


class FuelReportParser:
    """
    Robust parser for fuel report messages.
    Handles various formats, typos, and field ordering.
    """
    
    # Regex patterns for each field (case-insensitive, flexible spacing)
    PATTERNS = {
        'driver': [
            r'DRIVER\s*[:\-=]\s*(.+?)(?=\n|CAR|LITERS|LITRES|AMOUNT|TYPE|ODOMETER|$)',
            r'JINA\s*[:\-=]\s*(.+?)(?=\n|CAR|LITERS|LITRES|AMOUNT|TYPE|ODOMETER|$)',  # Swahili
            r'NAME\s*[:\-=]\s*(.+?)(?=\n|CAR|LITERS|LITRES|AMOUNT|TYPE|ODOMETER|$)',
        ],
        'car': [
            # Kenyan plates: 2-4 letters, 2-4 digits, optional trailing letter
            r'CAR\s*[:\-=]\s*([A-Z]{2,4}\s*\d{2,4}\s*[A-Z]?)(?=\s|$|\n|LITERS|LITRES|AMOUNT|TYPE|ODOMETER)',
            r'REG\s*(?:NO)?\.?\s*[:\-=]\s*([A-Z]{2,4}\s*\d{2,4}\s*[A-Z]?)(?=\s|$|\n)',
            r'VEHICLE\s*[:\-=]\s*([A-Z]{2,4}\s*\d{2,4}\s*[A-Z]?)(?=\s|$|\n)',
            r'PLATE\s*[:\-=]\s*([A-Z]{2,4}\s*\d{2,4}\s*[A-Z]?)(?=\s|$|\n)',
            r'GARI\s*[:\-=]\s*([A-Z]{2,4}\s*\d{2,4}\s*[A-Z]?)(?=\s|$|\n)',  # Swahili
            # Fallback: plate pattern anywhere in text
            r'\b([A-Z]{2,4}\s*\d{3,4}\s*[A-Z])\b',
        ],
        'liters': [
            r'LITERS?\s*[:\-=]\s*([\d,\.]+)',
            r'LITRES?\s*[:\-=]\s*([\d,\.]+)',
            r'LTR?S?\s*[:\-=]\s*([\d,\.]+)',
            r'FUEL\s*[:\-=]\s*([\d,\.]+)\s*(?:L|LTR)',
            r'([\d,\.]+)\s*(?:LITERS?|LITRES?|LTR?S?)\b',
        ],
        'amount': [
            r'AMOUNT\s*[:\-=]\s*(?:KSH?\.?\s*)?([\d,\.]+)',
            r'COST\s*[:\-=]\s*(?:KSH?\.?\s*)?([\d,\.]+)',
            r'PRICE\s*[:\-=]\s*(?:KSH?\.?\s*)?([\d,\.]+)',
            r'KSH?\.?\s*[:\-=]?\s*([\d,\.]+)',
            r'TOTAL\s*[:\-=]\s*(?:KSH?\.?\s*)?([\d,\.]+)',
            r'PESA\s*[:\-=]\s*([\d,\.]+)',  # Swahili
        ],
        'type': [
            r'TYPE\s*[:\-=]\s*(DIESEL|PETROL|SUPER|V-?POWER|UNLEADED|AGO)',
            r'FUEL\s*TYPE\s*[:\-=]\s*(DIESEL|PETROL|SUPER|V-?POWER|UNLEADED|AGO)',
            r'\b(DIESEL|PETROL|SUPER|V-?POWER|UNLEADED|AGO)\b',
        ],
        'odometer': [
            r'ODOMETER\s*[:\-=]\s*([\d,\.]+)',
            r'ODO\s*[:\-=]\s*([\d,\.]+)',
            r'KM\s*[:\-=]\s*([\d,\.]+)',
            r'MILEAGE\s*[:\-=]\s*([\d,\.]+)',
            r'READING\s*[:\-=]\s*([\d,\.]+)',
        ],
        'department': [
            r'DEPARTMENT\s*[:\-=]\s*(.+?)(?=\n|DRIVER|CAR|LITERS|LITRES|AMOUNT|TYPE|ODOMETER|$)',
            r'DEPT\s*[:\-=]\s*(.+?)(?=\n|DRIVER|CAR|LITERS|LITRES|AMOUNT|TYPE|ODOMETER|$)',
            r'SECTION\s*[:\-=]\s*(.+?)(?=\n|DRIVER|CAR|LITERS|LITRES|AMOUNT|TYPE|ODOMETER|$)',
        ],
    }
    
    # Valid fuel types (normalized)
    VALID_FUEL_TYPES = {'DIESEL', 'PETROL', 'SUPER', 'V-POWER', 'UNLEADED', 'AGO'}
    
    def __init__(self, config: Dict):
        self.config = config
        self.strict_mode = config.get('parsing', {}).get('strictMode', False)
        self.required_fields = config.get('parsing', {}).get('requiredFields', 
            ['driver', 'car', 'liters', 'amount', 'type', 'odometer'])
    
    def parse(self, message_body: str) -> Tuple[Optional[Dict], List[str]]:
        """
        Parse a fuel report message and extract structured data.
        
        Returns:
            Tuple of (parsed_data or None, list of errors/warnings)
        """
        errors = []
        data = {}
        
        # Normalize message: uppercase, clean whitespace
        text = message_body.upper().strip()
        text = re.sub(r'\s+', ' ', text)  # Normalize whitespace
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Also try with newlines preserved for pattern matching
        text_with_newlines = message_body.upper().strip()
        
        # Extract each field
        for field, patterns in self.PATTERNS.items():
            value = self._extract_field(text_with_newlines, patterns)
            if value is None:
                value = self._extract_field(text, patterns)
            
            if value:
                data[field] = self._clean_value(field, value)
            elif field in self.required_fields:
                errors.append(f"Missing required field: {field}")
        
        # Validate extracted data
        validation_errors = self._validate(data)
        errors.extend(validation_errors)
        
        # Determine success
        if self.strict_mode and errors:
            return None, errors
        elif not self.strict_mode and len(data) >= 2:
            # In lenient mode, accept if at least 2 fields found
            return data, errors
        elif len(errors) == len(self.required_fields):
            # No fields found at all
            return None, errors
        
        return data, errors
    
    def _extract_field(self, text: str, patterns: List[str]) -> Optional[str]:
        """Try each pattern until one matches."""
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                return match.group(1).strip()
        return None
    
    def _clean_value(self, field: str, value: str) -> Any:
        """Clean and convert extracted value to appropriate type."""
        value = value.strip()
        
        if field == 'driver':
            # Clean driver name: remove extra whitespace, capitalize
            return ' '.join(value.split()).title()
        
        elif field == 'department':
            # Clean department name: remove extra whitespace, uppercase
            return ' '.join(value.split()).upper()
        
        elif field == 'car':
            # Normalize plate: uppercase, standardize spacing
            plate = re.sub(r'\s+', ' ', value.upper()).strip()
            return plate
        
        elif field == 'liters':
            # Convert to float
            try:
                return float(value.replace(',', ''))
            except ValueError:
                return value
        
        elif field == 'amount':
            # Convert to float, remove commas
            try:
                return float(value.replace(',', ''))
            except ValueError:
                return value
        
        elif field == 'type':
            # Normalize fuel type
            fuel_type = value.upper().replace('-', '-').strip()
            if fuel_type == 'VPOWER':
                fuel_type = 'V-POWER'
            elif fuel_type == 'AGO':
                fuel_type = 'DIESEL'  # AGO is diesel
            return fuel_type
        
        elif field == 'odometer':
            # Convert to integer
            try:
                return int(float(value.replace(',', '')))
            except ValueError:
                return value
        
        return value
    
    def _validate(self, data: Dict) -> List[str]:
        """Validate extracted data."""
        errors = []
        
        # Validate liters
        if 'liters' in data:
            liters = data['liters']
            if isinstance(liters, (int, float)):
                if liters <= 0:
                    errors.append(f"Invalid liters value: {liters} (must be positive)")
                elif liters > 500:
                    errors.append(f"Warning: unusually high liters: {liters}")
        
        # Validate amount
        if 'amount' in data:
            amount = data['amount']
            if isinstance(amount, (int, float)):
                if amount <= 0:
                    errors.append(f"Invalid amount value: {amount} (must be positive)")
        
        # Validate fuel type
        if 'type' in data:
            fuel_type = data['type']
            if fuel_type not in self.VALID_FUEL_TYPES:
                errors.append(f"Unknown fuel type: {fuel_type}")
        
        # Validate odometer
        if 'odometer' in data:
            odo = data['odometer']
            if isinstance(odo, (int, float)):
                if odo <= 0:
                    errors.append(f"Invalid odometer value: {odo} (must be positive)")
        
        # Validate car plate format (Kenyan plates: KXX 000X or similar)
        if 'car' in data:
            plate = data['car']
            if not re.match(r'^[A-Z]{2,3}\s*\d{2,4}\s*[A-Z]{0,3}$', plate, re.IGNORECASE):
                errors.append(f"Warning: unusual plate format: {plate}")
        
        return errors


class ExcelExporter:
    """
    Handles Excel file creation and appending with backup mechanism.
    """
    
    COLUMNS = ['DATETIME', 'DEPARTMENT', 'DRIVER', 'CAR', 'LITERS', 'AMOUNT', 'TYPE', 'ODOMETER', 'SENDER', 'RAW_MESSAGE']
    BACKUP_INTERVAL_RECORDS = 50  # Create backup every N records
    
    def __init__(self, output_folder: str, filename: str):
        self.output_folder = Path(output_folder)
        self.filename = filename
        self.filepath = self.output_folder / filename
        self.backup_folder = self.output_folder / 'backups'
        self.records_since_backup = 0
        
        # Ensure output folder exists
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.backup_folder.mkdir(parents=True, exist_ok=True)
    
    def create_backup(self) -> bool:
        """Create a timestamped backup of the Excel file."""
        if not self.filepath.exists():
            return False
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"{self.filename.replace('.xlsx', '')}_{timestamp}.xlsx"
            backup_path = self.backup_folder / backup_name
            shutil.copy2(self.filepath, backup_path)
            logger.info(f"[BACKUP] Created backup: {backup_name}")
            
            # Clean old backups (keep last 10)
            backups = sorted(self.backup_folder.glob(f"{self.filename.replace('.xlsx', '')}_*.xlsx"))
            if len(backups) > 10:
                for old_backup in backups[:-10]:
                    try:
                        old_backup.unlink()
                    except:
                        pass
            
            return True
        except Exception as e:
            logger.error(f"Backup failed: {e}")
            return False
    
    def get_last_odometer_for_car(self, car_plate: str) -> Optional[int]:
        """Get the last recorded odometer reading for a specific car."""
        if not self.filepath.exists():
            return None
        
        try:
            wb = load_workbook(self.filepath, read_only=True)
            ws = wb.active
            
            # Normalize the plate for comparison
            car_plate_normalized = car_plate.upper().replace(' ', '')
            
            last_odometer = None
            
            # Skip header row, iterate through data
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 8:
                    row_car = str(row[3] or '').upper().replace(' ', '')  # CAR is column D (index 3)
                    row_odometer = row[7]  # ODOMETER is column H (index 7)
                    
                    if row_car == car_plate_normalized and row_odometer:
                        try:
                            odo_value = int(float(str(row_odometer).replace(',', '')))
                            last_odometer = odo_value
                        except (ValueError, TypeError):
                            continue
            
            wb.close()
            return last_odometer
            
        except Exception as e:
            logger.error(f"Error reading last odometer: {e}")
            return None
    
    def validate_odometer(self, car_plate: str, new_odometer: int, driver: str, sender_phone: str = '') -> Tuple[bool, Optional[str]]:
        """
        Validate that the new odometer reading is greater than the previous one.
        Returns (is_valid, error_message)
        """
        if not new_odometer:
            return True, None  # No odometer to validate
        
        last_odometer = self.get_last_odometer_for_car(car_plate)
        
        if last_odometer is None:
            return True, None  # First record for this car
        
        if new_odometer <= last_odometer:
            error_msg = f"Odometer {new_odometer} is not greater than previous reading {last_odometer}"
            logger.warning(f"[ERROR] Odometer validation failed for {car_plate}: {error_msg}")
            
            # Save validation error for WhatsApp notification
            save_validation_error(
                car=car_plate,
                driver=driver,
                issue=f"Odometer reading {new_odometer:,} km is less than or equal to previous reading {last_odometer:,} km. Please verify and resend.",
                sender_phone=sender_phone
            )
            
            return False, error_msg
        
        return True, None
    
    def append_record(self, record: Dict, validate_odometer: bool = True) -> Tuple[bool, Optional[str]]:
        """
        Append a single record to the Excel file.
        Returns (success, error_message)
        """
        try:
            car_plate = record.get('car', '')
            odometer = record.get('odometer')
            driver = record.get('driver', 'Unknown')
            sender_phone = record.get('sender_phone', '')
            
            # Validate odometer if enabled
            if validate_odometer and odometer and car_plate:
                try:
                    odo_int = int(float(str(odometer).replace(',', '')))
                    is_valid, error_msg = self.validate_odometer(car_plate, odo_int, driver, sender_phone)
                    if not is_valid:
                        return False, error_msg
                except (ValueError, TypeError):
                    pass  # Skip validation if odometer is not a valid number
            
            # Prepare row data
            row = {
                'DATETIME': record.get('datetime', ''),
                'DEPARTMENT': record.get('department', ''),
                'DRIVER': record.get('driver', ''),
                'CAR': record.get('car', ''),
                'LITERS': record.get('liters', ''),
                'AMOUNT': record.get('amount', ''),
                'TYPE': record.get('type', ''),
                'ODOMETER': record.get('odometer', ''),
                'SENDER': record.get('sender', ''),
                'RAW_MESSAGE': record.get('raw_message', '')[:500],  # Truncate long messages
            }
            
            if self.filepath.exists():
                # Load existing workbook
                wb = load_workbook(self.filepath)
                ws = wb.active
                
                # Append row
                ws.append([row[col] for col in self.COLUMNS])
            else:
                # Create new workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'Fuel Records'
                
                # Add header row
                ws.append(self.COLUMNS)
                
                # Add data row
                ws.append([row[col] for col in self.COLUMNS])
                
                # Set column widths
                column_widths = {
                    'A': 18,  # DATETIME
                    'B': 15,  # DEPARTMENT
                    'C': 15,  # DRIVER
                    'D': 12,  # CAR
                    'E': 10,  # LITERS
                    'F': 12,  # AMOUNT
                    'G': 10,  # TYPE
                    'H': 12,  # ODOMETER
                    'I': 15,  # SENDER
                    'J': 50,  # RAW_MESSAGE
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
            
            # Save workbook with retry logic
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    wb.save(self.filepath)
                    break
                except PermissionError:
                    if attempt < max_retries - 1:
                        import time
                        time.sleep(0.5)  # Wait and retry
                    else:
                        raise
            
            logger.info(f"Appended record to {self.filepath}")
            
            # Increment counter and backup if needed
            self.records_since_backup += 1
            if self.records_since_backup >= self.BACKUP_INTERVAL_RECORDS:
                self.create_backup()
                self.records_since_backup = 0
            
            return True, None
            
        except PermissionError as e:
            logger.error(f"Permission denied writing to Excel (file may be open): {e}")
            return False, "Excel file is open in another program. Please close it and try again."
        except Exception as e:
            logger.error(f"Error appending to Excel: {e}")
            return False, str(e)
    
    def append_records(self, records: List[Dict]) -> int:
        """Append multiple records. Returns count of successful appends."""
        success_count = 0
        for record in records:
            success, _ = self.append_record(record)
            if success:
                success_count += 1
        return success_count

    def update_record(self, original_datetime: str, original_car: str, new_record: Dict) -> Tuple[bool, Optional[str]]:
        """
        Update an existing record in the Excel file by datetime and car.
        Used for edit approvals.
        Returns (success, error_message)
        """
        try:
            if not self.filepath.exists():
                return False, "Excel file does not exist"
            
            wb = load_workbook(self.filepath)
            ws = wb.active
            
            # Normalize car plate for comparison
            original_car_normalized = original_car.upper().replace(' ', '')
            
            # Find the row to update (datetime is col A, car is col D)
            target_row = None
            for row_idx in range(2, ws.max_row + 1):  # Skip header
                row_datetime = str(ws.cell(row=row_idx, column=1).value or '')
                row_car = str(ws.cell(row=row_idx, column=4).value or '').upper().replace(' ', '')
                
                if row_datetime == original_datetime and row_car == original_car_normalized:
                    target_row = row_idx
                    break
            
            if not target_row:
                wb.close()
                return False, f"Record not found: {original_datetime} / {original_car}"
            
            # Prepare row data
            row_data = {
                'DATETIME': new_record.get('datetime', ''),
                'DEPARTMENT': new_record.get('department', ''),
                'DRIVER': new_record.get('driver', ''),
                'CAR': new_record.get('car', ''),
                'LITERS': new_record.get('liters', ''),
                'AMOUNT': new_record.get('amount', ''),
                'TYPE': new_record.get('type', ''),
                'ODOMETER': new_record.get('odometer', ''),
                'SENDER': new_record.get('sender', ''),
                'RAW_MESSAGE': new_record.get('raw_message', '')[:500],
            }
            
            # Update each cell in the row
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                ws.cell(row=target_row, column=col_idx, value=row_data[col_name])
            
            wb.save(self.filepath)
            wb.close()
            logger.info(f"Updated record in Excel: {original_datetime} / {original_car}")
            return True, None
            
        except Exception as e:
            logger.error(f"Error updating Excel record: {e}")
            return False, str(e)

def load_config() -> Dict:
    """Load configuration from config.json."""
    try:
        with open(CONFIG_PATH, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading config: {e}")
        return {}


def format_datetime(timestamp: int) -> str:
    """Format Unix timestamp to YYYY-MM-DD-HH-MM."""
    try:
        dt = datetime.fromtimestamp(timestamp)
        return dt.strftime('%Y-%m-%d-%H-%M')
    except:
        return datetime.now().strftime('%Y-%m-%d-%H-%M')


def process_message_file(filepath: Path, parser: FuelReportParser, exporter: ExcelExporter) -> bool:
    """
    Process a single message file.
    
    Returns True if successfully processed (or moved to errors), False if should retry.
    """
    try:
        with open(filepath, 'r') as f:
            message_data = json.load(f)
        
        body = message_data.get('body', '')
        
        if not body:
            logger.warning(f"Empty message body in {filepath.name}")
            move_to_errors(filepath, "Empty message body")
            return True
        
        # Parse the message
        parsed, errors = parser.parse(body)
        
        if parsed and len(parsed) >= 3:  # At least 3 fields extracted
            # Create record for Excel
            record = {
                'datetime': format_datetime(message_data.get('timestamp', 0)),
                'department': parsed.get('department', ''),
                'driver': parsed.get('driver', ''),
                'car': parsed.get('car', ''),
                'liters': parsed.get('liters', ''),
                'amount': parsed.get('amount', ''),
                'type': parsed.get('type', ''),
                'odometer': parsed.get('odometer', ''),
                'sender': message_data.get('senderName', ''),
                'sender_phone': message_data.get('senderPhone', ''),
                'raw_message': body,
            }

            # Validate ALL required fields are present and not empty
            required_fields = {
                'department': 'DEPARTMENT',
                'driver': 'DRIVER',
                'car': 'CAR/VEHICLE',
                'liters': 'LITERS',
                'amount': 'AMOUNT',
                'type': 'TYPE (DIESEL/PETROL)',
                'odometer': 'ODOMETER',
            }
            
            missing_fields = []
            for field, label in required_fields.items():
                value = record.get(field)
                if not value or (isinstance(value, str) and not value.strip()):
                    missing_fields.append(label)
            
            if missing_fields:
                issue = f"Missing required field(s): {', '.join(missing_fields)}"
                car_plate = record.get('car', 'UNKNOWN')
                driver = record.get('driver', 'Unknown')
                sender_phone = record.get('sender_phone', '')
                logger.warning(f"[DENIED] {issue} in {filepath.name}")
                save_validation_error(car_plate, driver, issue, sender_phone)
                move_to_errors(filepath, f"MISSING FIELDS: {', '.join(missing_fields)}")
                return True

            # Normalize the car plate for validation
            normalized_plate = normalize_plate(record['car'])
            
            # Validate car plate against allowed fleet list (strict match)
            if normalized_plate not in ALLOWED_PLATES:
                issue = f"Vehicle {record['car']} is not in the approved fleet list. Please check the registration number."
                logger.warning(f"[DENIED] Unauthorized plate: {record['car']} (normalized: {normalized_plate}) in {filepath.name}")
                save_validation_error(record['car'], record.get('driver', ''), issue, record.get('sender_phone', ''))
                move_to_errors(filepath, f"UNAUTHORIZED PLATE: {record['car']}")
                return True
            
            # Use normalized plate in record for consistency
            record['car'] = normalized_plate
            
            # Check if this is an admin-approved record (skip consistency checks)
            is_approved = message_data.get('isApproved', False)
            
            # Check car cooldown (same car can't fuel within 12 hours)
            # Skip this check for admin-approved records
            if not is_approved:
                cooldown_ok, cooldown_approval_id = check_car_cooldown(
                    normalized_plate,
                    record
                )
                
                if not cooldown_ok:
                    logger.warning(f"[PENDING] Car cooldown violation: {filepath.name} - ID: {cooldown_approval_id}")
                    move_to_errors(filepath, f"PENDING APPROVAL: {cooldown_approval_id} - Car cooldown")
                    return True
            else:
                logger.info(f"[OK] Admin-approved record - skipping cooldown check")
            
            # NOTE: Driver consistency check removed - the 12-hour car cooldown is the primary protection.
            # If same car fuels within 12 hours (with same or different driver), it requires approval.
            # After 12 hours, any driver can fuel the car normally.
            
            # Check if this is an EDIT approval (should UPDATE, not INSERT)
            is_edit_approval = message_data.get('approvalType') == 'edit'
            original_approval_id = message_data.get('originalApprovalId')
            
            # Get original record info for edit updates
            original_datetime = None
            original_car = None
            if is_edit_approval and original_approval_id:
                # Load original record from pending_approvals
                try:
                    if PENDING_APPROVALS_PATH.exists():
                        with open(PENDING_APPROVALS_PATH, 'r') as f:
                            approvals = json.load(f)
                        for approval in approvals:
                            if approval.get('id') == original_approval_id:
                                orig = approval.get('original_record', {})
                                original_datetime = orig.get('datetime', '')
                                original_car = orig.get('car', '')
                                break
                except Exception as e:
                    logger.error(f"Error loading original record for edit: {e}")
            
            if is_edit_approval and original_datetime and original_car:
                # UPDATE existing record instead of appending
                success, error_msg = exporter.update_record(original_datetime, original_car, record)
                if success:
                    logger.info(f"[EDIT] Updated record (edit approval): {original_car}")
                else:
                    # Fallback to append if update fails
                    logger.warning(f"[WARN] Update failed, appending instead: {error_msg}")
                    success, error_msg = exporter.append_record(record)
            else:
                # Normal append (new record or cooldown approval)
                success, error_msg = exporter.append_record(record)
            
            if success:
                # Move to processed folder
                move_to_processed(filepath)
                logger.info(f"[OK] Processed: {filepath.name} -> {parsed.get('car', 'N/A')}")
                
                # Calculate fuel efficiency
                efficiency = None
                distance = None
                alert = None
                try:
                    current_odo = int(float(str(record.get('odometer', 0)).replace(',', '')))
                    current_liters = float(str(record.get('liters', 0)).replace(',', ''))
                    if current_odo > 0 and current_liters > 0:
                        efficiency, alert = calculate_fuel_efficiency(
                            record.get('car', ''),
                            current_odo,
                            current_liters
                        )
                        if efficiency is not None:
                            # Get distance for confirmation message
                            last_update = get_car_last_update(record.get('car', ''))
                            if last_update:
                                prev_odo = int(float(str(last_update.get('odometer', 0)).replace(',', '')))
                                distance = current_odo - prev_odo
                                # Save efficiency record for history
                                save_efficiency_record(
                                    record.get('car', ''),
                                    efficiency,
                                    distance,
                                    float(str(last_update.get('liters', 0)).replace(',', '')),
                                    record.get('driver', '')
                                )
                                logger.info(f"[EFFICIENCY] {record.get('car', '')}: {efficiency:.1f} km/L over {distance:,} km")
                        # Send alert if needed
                        if alert:
                            save_efficiency_alert(record.get('car', ''), record.get('driver', ''), alert)
                            logger.warning(f"[ALERT] Efficiency alert for {record.get('car', '')}: {alert['type']}")
                except Exception as e:
                    logger.error(f"Error calculating efficiency: {e}")
                
                # Update car last update timestamp (for 12h cooldown - also stores efficiency)
                update_car_last_update(record.get('car', ''), record, efficiency)
                
                # Save confirmation for WhatsApp notification (with efficiency if available)
                save_confirmation(record, message_data.get('senderName', 'Unknown'), efficiency, distance)
                
                # Optional: upload to Google Sheets and/or Database
                try:
                    cfg = load_config().get('upload', {})
                    # Google Sheets
                    if cfg.get('toGoogleSheets'):
                        from env import get_env
                        from google_sheets_uploader import GoogleSheetsUploader
                        sheet_cfg = cfg.get('google', {})
                        spreadsheet_id = sheet_cfg.get('spreadsheetId') or get_env('GOOGLE_SHEETS_SPREADSHEET_ID')
                        sheet_name = sheet_cfg.get('sheetName', 'FUEL RECORDS')
                        spreadsheet_name_env = get_env('GOOGLE_SHEETS_SPREADSHEET_NAME')
                        uploader = GoogleSheetsUploader(
                            spreadsheet_id=spreadsheet_id,
                            spreadsheet_name=spreadsheet_name_env or 'Fuel Records',
                            worksheet_name=sheet_name,
                        )
                        uploader.ensure_headers(ExcelExporter.COLUMNS)
                        
                        # For edit approvals, UPDATE instead of append
                        if is_edit_approval and original_datetime and original_car:
                            if uploader.update_record(original_datetime, original_car, record, ExcelExporter.COLUMNS):
                                logger.info("📤 Updated record in Google Sheets")
                            else:
                                # Fallback to append if update fails
                                uploader.append_record(record, ExcelExporter.COLUMNS)
                                logger.info("📤 Appended record to Google Sheets (update not found)")
                        else:
                            uploader.append_record(record, ExcelExporter.COLUMNS)
                            logger.info("📤 Uploaded record to Google Sheets")
                except Exception as e:
                    logger.error(f"Google Sheets upload failed: {e}")

                try:
                    cfg = load_config().get('upload', {})
                    if cfg.get('toDatabase'):
                        from db import Database
                        table_name = (cfg.get('database') or {}).get('tableName', 'fuel_records')
                        db = Database(table_name=table_name)
                        
                        # For edit approvals, UPDATE instead of insert
                        if is_edit_approval and original_datetime and original_car:
                            if db.update_fuel_record(original_datetime, original_car, record):
                                logger.info("[DB] Updated record in database")
                            else:
                                # Fallback to insert if update fails
                                if db.insert_fuel_record(record):
                                    logger.info("[DB] Inserted record into database (update not found)")
                                else:
                                    logger.error("Database insert failed")
                        else:
                            if db.insert_fuel_record(record):
                                logger.info("[DB] Inserted record into database")
                            else:
                                logger.error("Database insert failed")
                except Exception as e:
                    logger.error(f"Database upload failed: {e}")

                if errors:
                    logger.warning(f"   Warnings: {', '.join(errors)}")
                
                return True
            elif error_msg and 'Odometer' in error_msg:
                # Odometer validation failed - move to errors but don't retry
                logger.warning(f"[WARN] Odometer validation failed: {filepath.name} - {error_msg}")
                move_to_errors(filepath, f"ODOMETER ERROR: {error_msg}")
                return True
            else:
                logger.error(f"Failed to save to Excel: {filepath.name}")
                return False
        else:
            # Could not parse enough fields
            error_msg = '; '.join(errors) if errors else 'Could not extract required fields'
            logger.warning(f"[ERROR] Parse failed: {filepath.name} - {error_msg}")
            move_to_errors(filepath, error_msg)
            return True
            
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in {filepath.name}: {e}")
        move_to_errors(filepath, f"Invalid JSON: {e}")
        return True
    except Exception as e:
        logger.error(f"Error processing {filepath.name}: {e}")
        return False


def move_to_processed(filepath: Path):
    """Move file to processed folder."""
    dest = PROCESSED_PATH / filepath.name
    try:
        PROCESSED_PATH.mkdir(parents=True, exist_ok=True)
        shutil.move(str(filepath), str(dest))
    except Exception as e:
        logger.error(f"Error moving to processed: {e}")


def move_to_errors(filepath: Path, error_msg: str):
    """Move file to errors folder with error info."""
    try:
        ERRORS_PATH.mkdir(parents=True, exist_ok=True)
        
        # Read original content
        with open(filepath, 'r') as f:
            content = json.load(f)
        
        # Add error info
        content['_parse_error'] = error_msg
        content['_error_time'] = datetime.now().isoformat()
        
        # Write to errors folder
        dest = ERRORS_PATH / filepath.name
        with open(dest, 'w') as f:
            json.dump(content, f, indent=2)
        
        # Remove original
        filepath.unlink()
        
    except Exception as e:
        logger.error(f"Error moving to errors folder: {e}")


def process_all_messages():
    """Process all pending message files."""
    logger.info("=" * 60)
    logger.info("Starting message processing...")
    
    # Load config
    config = load_config()
    if not config:
        logger.error("Could not load configuration. Exiting.")
        return
    
    # Initialize parser and exporter
    parser = FuelReportParser(config)
    
    output_folder = config.get('output', {}).get('excelFolder', './data/output')
    excel_filename = config.get('output', {}).get('excelFileName', 'fuel_records.xlsx')
    
    # Resolve relative path
    if output_folder.startswith('./'):
        output_folder = ROOT_DIR / output_folder[2:]
    
    exporter = ExcelExporter(output_folder, excel_filename)
    
    # Find all pending message files
    RAW_MESSAGES_PATH.mkdir(parents=True, exist_ok=True)
    message_files = list(RAW_MESSAGES_PATH.glob('msg_*.json'))
    
    if not message_files:
        logger.info("No new messages to process.")
        return
    
    logger.info(f"Found {len(message_files)} message(s) to process.")
    
    # Process each file
    success_count = 0
    error_count = 0
    
    for filepath in sorted(message_files):
        if process_message_file(filepath, parser, exporter):
            if (PROCESSED_PATH / filepath.name).exists():
                success_count += 1
            else:
                error_count += 1
    
    logger.info(f"Processing complete: {success_count} successful, {error_count} errors")
    logger.info("=" * 60)


def run_scheduler():
    """Run the processor on a schedule."""
    import schedule
    import time
    
    config = load_config()
    # Use seconds for quick response (default 10 seconds)
    interval_seconds = config.get('schedule', {}).get('processingIntervalSeconds', 10)
    
    logger.info(f"Starting scheduler - processing every {interval_seconds} seconds")
    logger.info("Press Ctrl+C to stop.\n")
    
    # Run immediately on start
    process_all_messages()
    
    # Schedule regular runs (every N seconds for quick feedback)
    schedule.every(interval_seconds).seconds.do(process_all_messages)
    
    try:
        while True:
            schedule.run_pending()
            time.sleep(10)
    except KeyboardInterrupt:
        logger.info("\nScheduler stopped by user.")


if __name__ == '__main__':
    import sys
    
    print("\n" + "=" * 60)
    print("  WhatsApp Fuel Extractor - Python Processor")
    print("=" * 60 + "\n")
    
    if len(sys.argv) > 1 and sys.argv[1] == '--once':
        # Run once and exit
        process_all_messages()
    else:
        # Run on schedule
        run_scheduler()
